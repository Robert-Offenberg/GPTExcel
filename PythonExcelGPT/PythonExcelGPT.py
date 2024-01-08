
from openpyxl import load_workbook

from openai import OpenAI

client = OpenAI()



def find_first_empty_row(worksheet, column_index=1):

    """

    Find the index of the first empty row in the specified column of a worksheet.

    

    Parameters:

    - worksheet: openpyxl worksheet object

    - column_index: index of the column to check for emptiness (default is 1 for column A)

    

    Returns:

    - Index of the first empty row in the specified column

    """

    for row_index in range(1, worksheet.max_row + 1):

        cell_value = worksheet.cell(row=row_index, column=column_index).value

        if cell_value is None or cell_value == '':

            return row_index

    

    # If no empty row is found, return the next row index after the last row

    return worksheet.max_row + 1



filename="output.xlsx"

gptmodel="gpt-3.5-turbo"

prompt=[

    {"role": "system", "content": "You are a Real Estate Agent"},

    {"role": "user", "content": "What is the best area to buy a house in Jacksonville, FL"}

  ]





completion = client.chat.completions.create(

  model=gptmodel,

  messages=prompt

)



workbook = load_workbook(filename)

worksheet = workbook.active

print(workbook.active.title)





empty_row_index = find_first_empty_row(worksheet)

print(completion.choices[0].message)

print(empty_row_index)

worksheet['A'+str(empty_row_index)] = str(prompt)

worksheet['B'+str(empty_row_index)] = str(completion.choices[0].message)

worksheet['C'+str(empty_row_index)] = gptmodel

workbook.save(filename)